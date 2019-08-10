
# coding: utf-8

# In[7]:

import subprocess
import sys

def install(package):
    subprocess.call([sys.executable, "-m", "pip", "install", package])

try:
    import pandas as pd
    import numpy as np
    import xlrd

except:
    install('pandas')
    install('numpy')
    install('xlrd')


import pandas as pd
import numpy as np
import os


# In[8]:


if not os.path.exists("Merged"):
    os.mkdir("Merged") 


# In[9]:


df1 = pd.read_excel("")
df2 = pd.read_excel("")


# In[10]:


merged = df1.merge(df2, how="", on="")
#Если в одном из файлов есть дубликаты, раскомментировать эту строку:
#merged = df1.merge(df2.drop_duplicates(subset=['A']), how="", on="")


# In[11]:


from datetime import datetime
thetime= datetime.now().strftime("%Y%m%d-%H%M%S")
merged.to_excel("Merged\\merge_"+thetime+".xlsx")


# In[12]:


"""

Без кавычек:

#%load file.py
%%writefile file.py  - в начале блока
%pycat  -
%run file.py
%lsmagic

from IPython.core.interactiveshell import InteractiveShell
InteractiveShell.ast_node_interactivity = "all"

===savetime===
from datetime import datetime
thetime= datetime.now().strftime("%Y%m%d-%H%M%S")

===openpyxl===
from openpyxl import load_workbook
from openpyxl import Workbook

wb = load_workbook()
wb_ws = wb.get_active_sheet()

wrwb = Workbook()
wrwb_ws = wrwb.get_active_sheet()

wb.save()

===numpy/pandas===
import pandas as pd
import numpy as np

excel_df = pd.read_excel()
csv_df = pd.read_csv()


df.to_excel()
df.to_csv()

writer = pd.ExcelWriter('',engine='xlsxwriter',options={})
df.to_excel(writer)
writer.save()


====requests/BeautifulSoup===
import requests
from bs4 import BeautifulSoup

page = requests.get("http://yandex.ru")
page.encoding = "windows-1251"
soup = BeautifulSoup(''.join(page.text), "html.parser\"),
soup.findAll("div")


===Files and directories===
import os
FileList = os.listdir()

#if not os.path.exists("Folder"):
#   os.mkdir("Folder") 

"""

